"""
data.py
"""
from openpyxl.reader.excel import load_workbook

class WebData:
   """
   This class is used to contain all the data that are required to perform the testing for the Swag Labs
   The test data are given from the DDF.xlsx Excel file.
   """


   def __init__(self):
       self.url = "https://opensource-demo.orangehrmlive.com/web/index.php/auth/login"
       self.dashboardURL = "https://opensource-demo.orangehrmlive.com/web/index.php/dashboard/index"
       self.fileName = "Data/DDF.xlsx"
       self.sheetName = "Sheet1"
       self.workbook = load_workbook(self.fileName)
       self.sheet = self.workbook[self.sheetName]


   def rowCount(self):
       """
       This method returns the maximum number of rows present in the Sheet 1
       :return: int
       """
       return self.sheet.max_row


   def readData(self, row, column):
       """
       This method will return the data present in the particular cell in the Sheet 1
       :param row:
       :param column:
       :return:
       """
       return self.sheet.cell(row, column).value


   def writeData(self, row, column, data):
       self.sheet.cell(row, column).value = data
       self.workbook.save(self.fileName)
